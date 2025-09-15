from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import openpyxl
import random

def get_words():
    # Fetch the webpage
    url = "https://www.ef.edu/english-resources/english-vocabulary/top-1000-words/"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    # Find the class 'field-item even'
    elements = soup.find_all(class_='field-item even')
    #Get paragraphs from the class
    paragraphs = elements[0].find_all('p')
    #extract words from the second paragraph
    second_para = paragraphs[1].get_text()
    words = second_para.split()

    #create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    #start at the first cell/ cell count
    count = 1

    #load words into different rows, but all in column one
    for word in words:
        index = f'A{count}'
        ws[index] = word.lower()
        count += 1

    #save the workbook after each word is input
    wb.save('New_Word_List.xlsx')


def game():
    #load/activate the workbook with the 1000 words
    wb = load_workbook('New_Word_List.xlsx')
    ws=wb.active

    #generate a random word from the 1000 different words inside the Excel sheet
    random_word = random.randint(1, 1000)
    word = list(ws.cell(row=random_word, column=1).value)
    word_len = len(word)

    #Placeholders for later use
    word_to_guess = []
    letters_guessed = []

    #Preset the _ for the word length into word_to_guess
    for length in range(0, word_len):
        word_to_guess.append('-')
    mistakes = 0

    while mistakes != 5:
        #Making sure the word has not been guessed already
        if word == word_to_guess:
            print("Congratulations! You guessed the word!")
            break

        else:

            while True:
                #displaying _ along with guessed letters
                for j in range(word_len):
                    print(word_to_guess[j], end='')

                guess = input('\nGuess a letter: ').lower()

                #make sure the letter is a single alphabetical character
                if guess.isalpha() and len(guess) == 1:
                    #Make sure the same letter cannot be guessed twice
                    if guess not in letters_guessed:
                        letters_guessed.append(guess)
                        correct_guess = 0
                        #looking for the letter in the array of letters of the random word
                        for i in range(0, word_len):
                            if guess == word[i]:
                                word_to_guess[i] = guess
                                correct_guess += 1
                        #display that a correct letter has been input
                        if correct_guess >=1:
                            print("You've guessed a correct letter!")
                            break
                        #tally/display mistakes
                        else:
                            mistakes += 1
                            print(f"You've guessed a wrong letter! Mistakes: {mistakes}")
                            break
                    #Throwing exceptions for previously guess letters
                    else:
                        print('You have guessed the letter already.')

                # Throwing exceptions for not character/more than 1 character
                else:
                    print("Please guess a single letter.")
    #If the loop stops because of 5 mistakes this will display
    if mistakes == 5:
        print("You lost, better luck next time!")

#game loop, so the game will stop when the user decides
get_words()
while True:
    y_n = input("Do you want to play the word game? (yes/no) ").lower()
    if y_n == 'yes':
        game()
    else:
        break


